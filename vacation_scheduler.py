#!/usr/bin/env python3
"""
Summer Vacation Scheduler for ATC Employees

This script optimizes vacation scheduling to maximize vacation days while
maintaining shift coverage requirements.
"""

import csv
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Dict, List, Set, Tuple, Optional
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Constants
DEFAULT_SHIFT_HOURS = 8.0  # Default assumption for shift duration when calculation fails


class Employee:
    """Represents an employee with their skills and working hour constraints."""

    def __init__(self, employee_id: str, name: str, skills: Set[str],
                 weekly_target_hours: int = 37, max_hours_per_week: int = 48):
        self.id = employee_id
        self.name = name
        self.skills = skills
        self.weekly_target_hours = weekly_target_hours
        self.max_hours_per_week = max_hours_per_week
        self.vacation_days = 0  # Track assigned vacation days

    def __repr__(self):
        return f"Employee({self.name}, {self.skills})"


class CoverageRequirement:
    """Represents a coverage requirement for a shift."""

    def __init__(self, day_type: str, shift_id: str, required: int, required_skill: str):
        self.day_type = day_type
        self.shift_id = shift_id
        self.required = required
        self.required_skill = required_skill

    def __repr__(self):
        return f"CoverageRequirement({self.day_type}, {self.shift_id}, {self.required}, {self.required_skill})"


class Shift:
    """Represents a shift type."""

    def __init__(self, shift_id: str, name: str, start: str, end: str, category: str):
        self.id = shift_id
        self.name = name
        self.start = start
        self.end = end
        self.category = category

        # Parse times to minutes for easier comparison
        start_parts = start.split(':')
        end_parts = end.split(':')
        self.start_minutes = int(start_parts[0]) * 60 + int(start_parts[1])
        self.end_minutes = int(end_parts[0]) * 60 + int(end_parts[1])

        # Handle overnight shifts
        if self.end_minutes < self.start_minutes:
            self.end_minutes += 24 * 60

    def __repr__(self):
        return f"Shift({self.name}, {self.start}-{self.end}, {self.category})"


def load_shifts(filepath: str) -> Dict[str, 'Shift']:
    """Load shift definitions from CSV file.

    Expected format:
    id,name,start,end,cat
    1,FD,07:00,15:15,Day

    Returns:
        Dict mapping shift name -> Shift object
    """
    shifts = {}
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                shift_id = row['id'].strip()
                name = row['name'].strip()
                start = row['start'].strip()
                end = row['end'].strip()
                category = row['cat'].strip()
                shifts[name] = Shift(shift_id, name, start, end, category)
    except FileNotFoundError:
        print(f"Error: File '{filepath}' not found.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error reading shifts file: {e}", file=sys.stderr)
        sys.exit(1)

    return shifts


def load_employees(filepath: str) -> List[Employee]:
    """Load employees from CSV file."""
    employees = []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                employee_id = row['id'].strip()
                name = row['name'].strip().strip('"')
                weekly_target_hours = int(row['weekly_target_hours'].strip())
                max_hours_per_week = int(row['max_hours_per_week'].strip())
                skills_str = row['skills'].strip().strip('"')
                skills = set(s.strip() for s in skills_str.split(';') if s.strip())
                employees.append(Employee(employee_id, name, skills, weekly_target_hours, max_hours_per_week))
    except FileNotFoundError:
        print(f"Error: File '{filepath}' not found.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error reading employees file: {e}", file=sys.stderr)
        sys.exit(1)

    return employees


def load_coverage(filepath: str) -> List[CoverageRequirement]:
    """Load coverage requirements from CSV file."""
    coverage = []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                day_type = row['type'].strip().strip('"')
                shift_id = row['shift_id'].strip().strip('"')
                required = int(row['required'].strip())
                required_skill = row['required_skills'].strip().strip('"')
                coverage.append(CoverageRequirement(day_type, shift_id, required, required_skill))
    except FileNotFoundError:
        print(f"Error: File '{filepath}' not found.", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error reading coverage file: {e}", file=sys.stderr)
        sys.exit(1)

    return coverage


def get_coverage_requirements_by_type(coverage: List[CoverageRequirement]) -> Dict[str, List[CoverageRequirement]]:
    """Organize coverage requirements by day type."""
    requirements = defaultdict(list)
    for req in coverage:
        requirements[req.day_type].append(req)
    return dict(requirements)


def calculate_min_employees_needed(requirements: List[CoverageRequirement]) -> Tuple[int, Dict[str, int]]:
    """Calculate minimum employees needed for a day type.

    Returns:
        Tuple of (total_positions, skill_requirements)
    """
    total_positions = sum(req.required for req in requirements)
    skill_requirements = defaultdict(int)

    for req in requirements:
        if req.required_skill != "None":
            skill_requirements[req.required_skill] += req.required

    return total_positions, dict(skill_requirements)


def can_cover_with_employees(
    employees_available: List[Employee],
    total_needed: int,
    skill_requirements: Dict[str, int]
) -> bool:
    """Check if available employees can meet coverage needs."""

    if len(employees_available) < total_needed:
        return False

    # Check each skill requirement
    for skill, required in skill_requirements.items():
        available_with_skill = sum(1 for emp in employees_available if skill in emp.skills)
        if available_with_skill < required:
            return False

    return True


def optimize_vacation_schedule(
    employees: List[Employee],
    coverage_weekday: List[CoverageRequirement],
    coverage_weekend: List[CoverageRequirement],
    start_date: datetime,
    num_weeks: int,
    target_days_per_employee: int = 21
) -> Dict[str, List[datetime]]:
    """Optimize vacation scheduling to maximize vacation days while maintaining coverage.

    Strategy:
    1. Calculate how many employees can be on vacation each day
    2. Use iterative fair allocation to distribute vacation days
    3. Prioritize employees with fewer vacation days assigned

    Args:
        employees: List of all employees
        coverage_weekday: Coverage requirements for weekdays
        coverage_weekend: Coverage requirements for weekends
        start_date: Start of vacation period
        num_weeks: Number of weeks in the vacation period
        target_days_per_employee: Target vacation days per employee (default 21)

    Returns:
        Dict mapping employee name -> list of vacation dates
    """
    # Calculate minimum employees needed per day type
    weekday_total, weekday_skills = calculate_min_employees_needed(coverage_weekday)
    weekend_total, weekend_skills = calculate_min_employees_needed(coverage_weekend)

    total_employees = len(employees)

    # Calculate max vacation capacity per day type
    max_vacation_weekday = total_employees - weekday_total
    max_vacation_weekend = total_employees - weekend_total

    print("\nVacation capacity:")
    print(f"  Weekdays: up to {max_vacation_weekday} employees on vacation (need {weekday_total} working)")
    print(f"  Weekends: up to {max_vacation_weekend} employees on vacation (need {weekend_total} working)")

    # Generate dates for the period
    dates = []
    current = start_date
    end_date = start_date + timedelta(weeks=num_weeks)
    while current < end_date:
        dates.append(current)
        current += timedelta(days=1)

    # Count weekdays vs weekends
    weekdays = [d for d in dates if d.weekday() < 5]
    weekends = [d for d in dates if d.weekday() >= 5]

    # Calculate theoretical maximum vacation days possible
    total_vacation_capacity = max_vacation_weekday * len(weekdays) + max_vacation_weekend * len(weekends)
    theoretical_max_per_employee = total_vacation_capacity // total_employees

    print("\nTheoretical maximum:")
    print(f"  Total vacation-day capacity: {total_vacation_capacity}")
    print(f"  If distributed equally: {theoretical_max_per_employee} days per employee")

    if target_days_per_employee > theoretical_max_per_employee:
        print(
            f"\n  WARNING: Target of {target_days_per_employee} days exceeds "
            f"theoretical max of {theoretical_max_per_employee} days!")
        print("  Consider extending the vacation period or reducing the target.")

    print("\nAllocating consecutive vacation blocks with equal distribution...")
    print("  Strategy: Split employees into two equal groups")

    # Initialize vacation assignments
    vacation_schedule = {emp.name: [] for emp in employees}

    # Modified allocation algorithm: Split employees into TWO EQUAL GROUPS
    # Group 1 takes vacation in first half, Group 2 in second half
    import random
    random.seed(42)  # For reproducibility

    # Balance groups by weekly_target_hours
    # Sort employees by weekly_target_hours descending to enable balanced distribution
    employees_by_hours = sorted(employees, key=lambda e: e.weekly_target_hours, reverse=True)

    # Use alternating assignment (highest to group1, next highest to group2, etc.)
    # to balance total hours between groups
    group1_base = []
    group2_base = []
    for i, emp in enumerate(employees_by_hours):
        if i % 2 == 0:
            group1_base.append(emp)
        else:
            group2_base.append(emp)

    # With odd number of employees, one group will have one more person
    # No adjustment needed - the alternating pattern already handles this optimally

    total_hours_group1 = sum(e.weekly_target_hours for e in group1_base)
    total_hours_group2 = sum(e.weekly_target_hours for e in group2_base)

    print(f"  Group 1: {len(group1_base)} employees (total weekly hours: {total_hours_group1})")
    print(f"  Group 2: {len(group2_base)} employees (total weekly hours: {total_hours_group2})")

    # Calculate midpoint and block sizes once (used in main loop and fallback)
    mid_point = len(dates) // 2
    max_block_first_half = mid_point
    max_block_second_half = len(dates) - mid_point

    best_schedule = None
    best_min_days = 0
    best_max_spread = float('inf')  # Difference between max and min days
    best_total_days = 0

    # Try many different employee orderings to find equal distribution
    for attempt in range(20):  # Increased attempts for better equality
        temp_schedule = {emp.name: [] for emp in employees}
        temp_vacation_by_date = {date: set() for date in dates}

        # Different ordering strategies, but always use balanced groups
        if attempt == 0:
            group1 = sorted(group1_base, key=lambda e: e.name)
            group2 = sorted(group2_base, key=lambda e: e.name)
        elif attempt == 1:
            group1 = sorted(group1_base, key=lambda e: e.name, reverse=True)
            group2 = sorted(group2_base, key=lambda e: e.name, reverse=True)
        else:
            # Shuffle within each group to try different orderings
            group1 = list(group1_base)
            group2 = list(group2_base)
            random.seed(42 + attempt)
            random.shuffle(group1)
            random.shuffle(group2)

        # Try to allocate equal blocks to all employees
        # Start with largest block size that fits in the smaller half
        max_block_size = min(max_block_first_half, max_block_second_half)

        # Ensure we have a valid range - start from max possible and go down to minimum of 6 days
        start_size = min(target_days_per_employee, max_block_size)
        end_size = max(1, max_block_size - 8)  # At least try down to 1 day

        # Only proceed if we can fit at least some vacation days
        if start_size < 1:
            continue

        for target_block_size in range(start_size, end_size - 1, -1):
            # Skip if block won't fit in either half
            if target_block_size > mid_point or target_block_size > (len(dates) - mid_point):
                continue

            temp_schedule = {emp.name: [] for emp in employees}
            temp_vacation_by_date = {date: set() for date in dates}

            # Process Group 1 first (first half of period)
            for emp in group1:
                # Find a consecutive block of exactly target_block_size in FIRST HALF
                best_block = None

                # Try different start positions in first half
                # Ensure we don't go past the boundary
                max_start_first = max(0, mid_point - target_block_size + 1)
                for start_idx in range(0, max_start_first):
                    end_idx = start_idx + target_block_size
                    candidate_block = dates[start_idx:end_idx]

                    # Check if this employee can take vacation on all these days
                    can_take_all = True
                    for date in candidate_block:
                        if emp.name in temp_vacation_by_date[date]:
                            can_take_all = False
                            break

                        is_weekend = date.weekday() in (5, 6)
                        max_vacation_today = max_vacation_weekend if is_weekend else max_vacation_weekday
                        requirements = coverage_weekend if is_weekend else coverage_weekday
                        _, skill_requirements = calculate_min_employees_needed(requirements)

                        current_vacation_count = len(temp_vacation_by_date[date])
                        if current_vacation_count >= max_vacation_today:
                            can_take_all = False
                            break

                        employees_working = [e for e in employees
                                             if e.name not in temp_vacation_by_date[date] and e.name != emp.name]

                        total_needed = weekday_total if not is_weekend else weekend_total

                        if not can_cover_with_employees(employees_working, total_needed, skill_requirements):
                            can_take_all = False
                            break

                    if can_take_all:
                        best_block = candidate_block
                        break  # Take the first available block

                # Assign the block found to this employee
                if best_block:
                    for date in best_block:
                        temp_schedule[emp.name].append(date)
                        temp_vacation_by_date[date].add(emp.name)

            # Process Group 2 (second half of period)
            for emp in group2:
                # Find a consecutive block of exactly target_block_size in SECOND HALF
                best_block = None

                # Try different start positions in second half
                # Ensure we don't go past the end
                max_start_second = max(mid_point, len(dates) - target_block_size + 1)
                for start_idx in range(mid_point, max_start_second):
                    end_idx = start_idx + target_block_size
                    candidate_block = dates[start_idx:end_idx]

                    # Check if this employee can take vacation on all these days
                    can_take_all = True
                    for date in candidate_block:
                        if emp.name in temp_vacation_by_date[date]:
                            can_take_all = False
                            break

                        is_weekend = date.weekday() in (5, 6)
                        max_vacation_today = max_vacation_weekend if is_weekend else max_vacation_weekday
                        requirements = coverage_weekend if is_weekend else coverage_weekday
                        _, skill_requirements = calculate_min_employees_needed(requirements)

                        current_vacation_count = len(temp_vacation_by_date[date])
                        if current_vacation_count >= max_vacation_today:
                            can_take_all = False
                            break

                        employees_working = [e for e in employees
                                             if e.name not in temp_vacation_by_date[date] and e.name != emp.name]

                        total_needed = weekday_total if not is_weekend else weekend_total

                        if not can_cover_with_employees(employees_working, total_needed, skill_requirements):
                            can_take_all = False
                            break

                    if can_take_all:
                        best_block = candidate_block
                        break  # Take the first available block

                # Assign the block found to this employee
                if best_block:
                    for date in best_block:
                        temp_schedule[emp.name].append(date)
                        temp_vacation_by_date[date].add(emp.name)

            # Check if all employees got the same amount (or within 1 day)
            vacation_counts = [len(days) for days in temp_schedule.values()]
            if vacation_counts:
                min_days = min(vacation_counts)
                max_days = max(vacation_counts)
                spread = max_days - min_days

                # If spread is within tolerance, this is a candidate solution
                if spread <= 1:
                    # Evaluate this schedule
                    total_days = sum(vacation_counts)

                    # Keep the best schedule (prioritize low spread, then high min_days, then total_days)
                    if (spread < best_max_spread or
                        (spread == best_max_spread and min_days > best_min_days) or
                        (spread == best_max_spread and min_days == best_min_days and
                         total_days > best_total_days)):
                        best_max_spread = spread
                        best_min_days = min_days
                        best_total_days = total_days
                        best_schedule = temp_schedule

                    break  # Found a good solution for this attempt, move to next ordering

    # If we couldn't achieve equal distribution, fall back to best effort
    if best_schedule is None:
        print("  Warning: Could not achieve equal distribution within 1 day. Using best effort allocation.")
        temp_schedule = {emp.name: [] for emp in employees}
        temp_vacation_by_date = {date: set() for date in dates}

        # Use the balanced groups from above
        group1 = sorted(group1_base, key=lambda e: e.name)
        group2 = sorted(group2_base, key=lambda e: e.name)

        # Group 1 - first half
        for emp in group1:
            best_block = None
            for block_length in range(target_days_per_employee, 6, -1):
                for start_idx in range(0, mid_point - block_length + 1):
                    end_idx = start_idx + block_length
                    candidate_block = dates[start_idx:end_idx]

                    can_take_all = True
                    for date in candidate_block:
                        if emp.name in temp_vacation_by_date[date]:
                            can_take_all = False
                            break

                        is_weekend = date.weekday() in (5, 6)
                        max_vacation_today = max_vacation_weekend if is_weekend else max_vacation_weekday
                        requirements = coverage_weekend if is_weekend else coverage_weekday
                        _, skill_requirements = calculate_min_employees_needed(requirements)

                        current_vacation_count = len(temp_vacation_by_date[date])
                        if current_vacation_count >= max_vacation_today:
                            can_take_all = False
                            break

                        employees_working = [e for e in employees
                                             if e.name not in temp_vacation_by_date[date] and e.name != emp.name]

                        total_needed = weekday_total if not is_weekend else weekend_total

                        if not can_cover_with_employees(employees_working, total_needed, skill_requirements):
                            can_take_all = False
                            break

                    if can_take_all:
                        best_block = candidate_block
                        break
                if best_block:
                    break

            if best_block:
                for date in best_block:
                    temp_schedule[emp.name].append(date)
                    temp_vacation_by_date[date].add(emp.name)

        # Group 2 - second half
        for emp in group2:
            best_block = None
            for block_length in range(target_days_per_employee, 6, -1):
                for start_idx in range(mid_point, len(dates) - block_length + 1):
                    end_idx = start_idx + block_length
                    candidate_block = dates[start_idx:end_idx]

                    can_take_all = True
                    for date in candidate_block:
                        if emp.name in temp_vacation_by_date[date]:
                            can_take_all = False
                            break

                        is_weekend = date.weekday() in (5, 6)
                        max_vacation_today = max_vacation_weekend if is_weekend else max_vacation_weekday
                        requirements = coverage_weekend if is_weekend else coverage_weekday
                        _, skill_requirements = calculate_min_employees_needed(requirements)

                        current_vacation_count = len(temp_vacation_by_date[date])
                        if current_vacation_count >= max_vacation_today:
                            can_take_all = False
                            break

                        employees_working = [e for e in employees
                                             if e.name not in temp_vacation_by_date[date] and e.name != emp.name]

                        total_needed = weekday_total if not is_weekend else weekend_total

                        if not can_cover_with_employees(employees_working, total_needed, skill_requirements):
                            can_take_all = False
                            break

                    if can_take_all:
                        best_block = candidate_block
                        break
                if best_block:
                    break

            if best_block:
                for date in best_block:
                    temp_schedule[emp.name].append(date)
                    temp_vacation_by_date[date].add(emp.name)

        best_schedule = temp_schedule

    vacation_schedule = best_schedule

    # Analyze the group split
    if vacation_schedule:
        print("\n  Analyzing vacation groups...")

        # Calculate start dates for each employee
        emp_start_dates = {}
        for emp_name, vac_dates in vacation_schedule.items():
            if vac_dates:
                emp_start_dates[emp_name] = min(vac_dates)

        # Sort by start date
        sorted_by_start = sorted(emp_start_dates.items(), key=lambda x: x[1])

        mid_date_idx = len(dates) // 2
        mid_date = dates[mid_date_idx]

        group1_actual = [name for name, start in sorted_by_start if start < mid_date]
        group2_actual = [name for name, start in sorted_by_start if start >= mid_date]

        print(f"  Group 1 (vacation in first half): {len(group1_actual)} employees")
        print(f"  Group 2 (vacation in second half): {len(group2_actual)} employees")
        if sorted_by_start:
            print(f"  First vacation starts: {sorted_by_start[0][1].strftime('%Y-%m-%d')}")
            print(f"  Last vacation starts: {sorted_by_start[-1][1].strftime('%Y-%m-%d')}")

    return vacation_schedule


def print_vacation_results(
    vacation_schedule: Dict[str, List[datetime]],
    employees: List[Employee],
    num_weeks: int,
    target_days: int
):
    """Print vacation scheduling results."""

    print("\n" + "=" * 70)
    print("VACATION SCHEDULING RESULTS")
    print("=" * 70)

    vacation_counts = {name: len(dates) for name, dates in vacation_schedule.items()}

    total_days = num_weeks * 7
    total_vacation_days = sum(vacation_counts.values())
    avg_days = total_vacation_days / len(employees) if employees else 0
    min_days = min(vacation_counts.values()) if vacation_counts else 0
    max_days = max(vacation_counts.values()) if vacation_counts else 0

    print(f"\nPeriod: {num_weeks} weeks ({total_days} days)")
    print(f"Target vacation days per employee: {target_days}")
    print("\nResults:")
    print(f"  Total vacation days allocated: {total_vacation_days}")
    print(f"  Average per employee: {avg_days:.1f} days")
    print(f"  Minimum: {min_days} days")
    print(f"  Maximum: {max_days} days")

    # Distribution analysis
    at_target = sum(1 for v in vacation_counts.values() if v >= target_days)
    below_target = len(employees) - at_target

    print("\nDistribution:")
    print(f"  Employees at/above target ({target_days}+ days): {at_target}")
    print(f"  Employees below target: {below_target}")

    # Show per-employee summary
    print("\nPer-employee allocation:")
    print("-" * 70)

    for name in sorted(vacation_counts.keys()):
        days = vacation_counts[name]
        percentage = (days / total_days * 100) if total_days > 0 else 0
        status = "✓" if days >= target_days else "✗"
        print(f"  {status} {name:20s}: {days:3d} days ({percentage:5.1f}%)")

    print("=" * 70)


def calculate_shift_hours(shift_id: str, shifts: Dict[str, 'Shift']) -> float:
    """Calculate duration of a shift in hours."""
    shift = shifts.get(shift_id)
    if not shift:
        return DEFAULT_SHIFT_HOURS

    # Parse time strings
    try:
        start_parts = shift.start.split(':')
        end_parts = shift.end.split(':')
        start_h, start_m = int(start_parts[0]), int(start_parts[1])
        end_h, end_m = int(end_parts[0]), int(end_parts[1])

        start_mins = start_h * 60 + start_m
        end_mins = end_h * 60 + end_m

        # Handle overnight shifts
        if end_mins <= start_mins:
            end_mins += 24 * 60

        duration_mins = end_mins - start_mins
        return duration_mins / 60.0
    except (ValueError, IndexError, AttributeError):
        return DEFAULT_SHIFT_HOURS  # Default on error


def rebalance_shift_assignments(
    shift_assignments: Dict[str, Dict[datetime, str]],
    employees: List[Employee],
    vacation_dates_by_employee: Dict[str, Set[datetime]],
    coverage_weekday: List[CoverageRequirement],
    coverage_weekend: List[CoverageRequirement],
    dates: List[datetime],
    shifts: Dict[str, 'Shift'],
    employee_by_name: Dict[str, Employee]
) -> Dict[str, Dict[datetime, str]]:
    """Rebalance shift assignments to improve fairness.

    This function performs multiple passes to redistribute shifts from over-assigned
    employees to under-assigned employees while maintaining coverage requirements.

    Args:
        shift_assignments: Current shift assignments
        employees: List of all employees
        vacation_dates_by_employee: Set of vacation dates per employee
        coverage_weekday: Coverage requirements for weekdays
        coverage_weekend: Coverage requirements for weekends
        dates: List of dates in the period
        shifts: Dict mapping shift ID to Shift object
        employee_by_name: Dict mapping employee name to Employee object

    Returns:
        Rebalanced shift assignments
    """
    # Calculate current shift counts and hours
    shift_counts = defaultdict(int)
    total_hours = defaultdict(float)

    for emp_name, assignments in shift_assignments.items():
        for date, shift_id in assignments.items():
            shift_counts[emp_name] += 1
            total_hours[emp_name] += calculate_shift_hours(shift_id, shifts)

    # Get working employees (those not on vacation entire period)
    working_employees = [emp for emp in employees if shift_counts[emp.name] > 0]

    if not working_employees:
        return shift_assignments

    # Calculate target shift count (average)
    total_shifts = sum(shift_counts.values())
    avg_shifts = total_shifts / len(working_employees)

    # Calculate ideal min/max bounds (tight tolerance for fairness)
    min_target = int(avg_shifts) - 1  # Allow 1 below average
    max_target = int(avg_shifts) + 2  # Allow 2 above average

    initial_min = min(shift_counts[emp.name] for emp in working_employees)
    initial_max = max(shift_counts[emp.name] for emp in working_employees)
    initial_spread = initial_max - initial_min

    # Perform multiple rebalancing passes with progressively tighter constraints
    max_passes = 30  # Increased passes for more thorough balancing
    total_transfers = 0

    for pass_num in range(max_passes):
        # Find over-assigned and under-assigned employees
        # Use more aggressive thresholds as passes progress
        threshold_adjustment = max(0, 2 - (pass_num // 10))  # Start aggressive, become more lenient
        current_max_target = max_target + threshold_adjustment
        current_min_target = min_target - threshold_adjustment

        over_assigned = []
        under_assigned = []

        for emp in working_employees:
            count = shift_counts[emp.name]
            if count > current_max_target:
                over_assigned.append((emp, count - current_max_target))
            elif count < current_min_target:
                under_assigned.append((emp, current_min_target - count))

        # Sort by deviation magnitude
        over_assigned.sort(key=lambda x: x[1], reverse=True)
        under_assigned.sort(key=lambda x: x[1], reverse=True)

        if not over_assigned or not under_assigned:
            break  # Balanced enough

        # Try to transfer shifts
        transfers_made = 0
        for over_emp, _ in over_assigned:
            # Get shifts assigned to this employee
            over_emp_shifts = [(date, shift_id) for date, shift_id
                               in shift_assignments[over_emp.name].items()]

            # Shuffle shifts to try different ones each pass
            if pass_num > 0:
                import random
                random.seed(42 + pass_num)
                random.shuffle(over_emp_shifts)

            # Try to transfer some shifts to under-assigned employees
            for date, shift_id in over_emp_shifts:
                # Check if we can find a replacement
                is_weekend = date.weekday() >= 5
                requirements = coverage_weekend if is_weekend else coverage_weekday

                # Get requirements for this shift
                shift_reqs = [req for req in requirements if req.shift_id == shift_id]
                if not shift_reqs:
                    continue

                # Find skill requirements
                skill_needed = None
                for req in shift_reqs:
                    if req.required_skill != "None":
                        skill_needed = req.required_skill
                        break

                # Try under-assigned employees as replacements
                for under_emp, _ in under_assigned:
                    # Check if employee is available (not on vacation)
                    if date in vacation_dates_by_employee.get(under_emp.name, set()):
                        continue

                    # Check if employee already has a shift this day
                    if date in shift_assignments[under_emp.name]:
                        continue

                    # Check if employee has required skill (but all employees have all skills now)
                    if skill_needed and skill_needed not in under_emp.skills:
                        continue

                    # Check weekly hours constraint against both target and max
                    week_start = date - timedelta(days=date.weekday())
                    week_dates = [d for d in dates if (d - timedelta(days=d.weekday())) == week_start]

                    current_week_hours = sum(
                        calculate_shift_hours(shift_assignments[under_emp.name].get(d, ''), shifts)
                        for d in week_dates if d in shift_assignments[under_emp.name]
                    )

                    shift_hours = calculate_shift_hours(shift_id, shifts)

                    # Strict enforcement: never exceed max_hours_per_week
                    if current_week_hours + shift_hours > under_emp.max_hours_per_week:
                        continue

                    # Prefer not to exceed weekly_target_hours, but allow if necessary for fairness in later passes
                    if pass_num < 20:  # First 20 passes: respect target hours
                        if current_week_hours + shift_hours > under_emp.weekly_target_hours:
                            continue

                    # Transfer the shift
                    del shift_assignments[over_emp.name][date]
                    shift_assignments[under_emp.name][date] = shift_id

                    # Update counts
                    shift_counts[over_emp.name] -= 1
                    shift_counts[under_emp.name] += 1
                    total_hours[over_emp.name] -= shift_hours
                    total_hours[under_emp.name] += shift_hours

                    transfers_made += 1
                    total_transfers += 1
                    break  # Move to next shift

                # Stop if over-assigned employee is now balanced
                if shift_counts[over_emp.name] <= current_max_target:
                    break

        if transfers_made == 0:
            break  # No more improvements possible

    final_min = min(shift_counts[emp.name] for emp in working_employees)
    final_max = max(shift_counts[emp.name] for emp in working_employees)
    final_spread = final_max - final_min

    print(f"  Rebalancing completed: {total_transfers} shift transfers")
    print(f"  Shift count spread: {initial_spread} → {final_spread} (improved by {initial_spread - final_spread})")

    return shift_assignments


def assign_shifts_to_employees(
    employees: List[Employee],
    vacation_schedule: Dict[str, List[datetime]],
    coverage_weekday: List[CoverageRequirement],
    coverage_weekend: List[CoverageRequirement],
    dates: List[datetime],
    shifts: Dict[str, 'Shift']
) -> Dict[str, Dict[datetime, str]]:
    """Assign shifts to employees for each day based on requirements.

    Uses a fair distribution algorithm that:
    - Tracks working hours per employee over rolling 6-week periods
    - Respects weekly_target_hours and max_hours_per_week constraints
    - Distributes shifts fairly across all working employees
    - Prioritizes employees with fewer hours worked

    Args:
        employees: List of all employees
        vacation_schedule: Dict mapping employee name -> list of vacation dates
        coverage_weekday: Coverage requirements for weekdays
        coverage_weekend: Coverage requirements for weekends
        dates: List of dates in the period
        shifts: Dict mapping shift ID to Shift object

    Returns:
        Dict mapping employee name -> Dict mapping date -> shift assignment
    """
    # Helper functions for candidate filtering and sorting
    def get_valid_candidates_tiered(available_employees, week_start, shift_hours,
                                    assigned_today, current_date, required_skill=None):
        """Filter employees by tier to ensure coverage while respecting constraints.

        Returns candidates in three tiers:
        - Tier 1 (strict): Below target hours, under max, <6 consecutive days
        - Tier 2 (moderate): Under max hours and <6 days, may exceed target
        - Tier 3 (emergency): Any available employee (coverage takes priority)

        Args:
            available_employees: List of employees not on vacation
            week_start: Start of the week for hours calculation
            shift_hours: Hours for this shift
            assigned_today: Set of employees already assigned today
            current_date: Current date
            required_skill: Required skill (or None for any skill)

        Returns:
            Tuple of (tier1_candidates, tier2_candidates, tier3_candidates)
        """
        tier1 = []  # Strict: below target, under max, <6 consecutive days
        tier2 = []  # Moderate: under max and <6 days (may exceed target)
        tier3 = []  # Emergency: any available employee

        for emp in available_employees:
            if emp.name in assigned_today:
                continue
            if required_skill and required_skill not in emp.skills:
                continue

            week_hours = hours_per_week[(emp.name, week_start)]
            consecutive_days = consecutive_work_days.get(emp.name, 0)

            # Tier 3: Always available (coverage priority)
            tier3.append(emp)

            # Check if under max hours and consecutive days
            under_max_hours = (week_hours + shift_hours) <= emp.max_hours_per_week
            under_consecutive_limit = consecutive_days < 6

            if under_max_hours and under_consecutive_limit:
                # Check if also under target hours
                under_target_hours = (week_hours + shift_hours) <= emp.weekly_target_hours
                if under_target_hours:
                    tier1.append(emp)  # Tier 1: all constraints met
                else:
                    tier2.append(emp)  # Tier 2: meets hard limits, exceeds soft target

        return tier1, tier2, tier3

    def create_sort_key_for_employee(emp, week_start, shift_hours):
        """Create a sort key for fair employee assignment."""
        week_hours = hours_per_week[(emp.name, week_start)]
        would_exceed_target = (week_hours + shift_hours) > emp.weekly_target_hours
        return (would_exceed_target, week_hours, shift_counts[emp.name], total_hours[emp.name], emp.name)

    # Initialize shift assignments
    shift_assignments = {emp.name: {} for emp in employees}

    # Convert vacation schedule to sets for faster lookup
    vacation_dates_by_employee = {
        name: set(dates_list) for name, dates_list in vacation_schedule.items()
    }

    # Create employee lookup by name
    employee_by_name = {emp.name: emp for emp in employees}

    # Track hours worked per employee per week (using 6-week rolling window)
    # Key: (employee_name, week_start_date), Value: hours
    hours_per_week = defaultdict(float)

    # Track total hours per employee
    total_hours = defaultdict(float)

    # Track number of shifts per employee (for fairness)
    shift_counts = defaultdict(int)

    # Track consecutive working days per employee
    consecutive_work_days = defaultdict(int)
    last_work_date = {}

    # For each date, assign shifts
    for date in dates:
        is_weekend = date.weekday() >= 5
        requirements = coverage_weekend if is_weekend else coverage_weekday

        # Calculate which week this date falls into (Monday-Sunday weeks)
        week_start = date - timedelta(days=date.weekday())

        # Get employees available on this date
        employees_available = [emp for emp in employees
                               if date not in vacation_dates_by_employee.get(emp.name, set())]

        # Group requirements by shift
        shift_reqs = defaultdict(list)
        for req in requirements:
            shift_reqs[req.shift_id].append(req)

        # Track which employees are already assigned today
        assigned_today = set()

        # Assign employees to shifts using fair distribution algorithm
        for shift_id in sorted(shift_reqs.keys()):
            reqs = shift_reqs[shift_id]

            # Calculate total needed for this shift
            total_needed = sum(req.required for req in reqs)

            # Get skill requirements
            skill_needs = {}
            for req in reqs:
                if req.required_skill != "None":
                    skill_needs[req.required_skill] = req.required

            # Calculate shift duration in hours
            shift_hours = calculate_shift_hours(shift_id, shifts)

            # Build candidate lists for each skill requirement
            assigned_this_shift = []

            # First, assign employees with required skills
            for skill, needed in skill_needs.items():
                # Get valid candidates with required skill in tiers
                tier1, tier2, tier3 = get_valid_candidates_tiered(
                    employees_available, week_start, shift_hours, assigned_today, date, skill)

                # Sort each tier by fairness priority (prefer those below target hours)
                tier1.sort(key=lambda emp: create_sort_key_for_employee(emp, week_start, shift_hours))
                tier2.sort(key=lambda emp: create_sort_key_for_employee(emp, week_start, shift_hours))
                tier3.sort(key=lambda emp: create_sort_key_for_employee(emp, week_start, shift_hours))

                # Try to fill from tier 1 first, then tier 2, then tier 3
                candidates = tier1 + tier2 + tier3

                # Assign the needed number of employees (MUST fill all positions)
                assigned_count = 0
                used_tier = 1
                for i, emp in enumerate(candidates):
                    if assigned_count >= needed:
                        break

                    # Track which tier we're using
                    if i >= len(tier1) + len(tier2):
                        if used_tier < 3:
                            print(
                                f"  WARNING: Using emergency tier for {shift_id} "
                                f"on {date.strftime('%Y-%m-%d')} (skill: {skill})")
                            used_tier = 3
                    elif i >= len(tier1):
                        used_tier = 2

                    shift_assignments[emp.name][date] = shift_id
                    assigned_today.add(emp.name)
                    assigned_this_shift.append(emp.name)
                    assigned_count += 1

                    # Update tracking
                    hours_per_week[(emp.name, week_start)] += shift_hours
                    total_hours[emp.name] += shift_hours
                    shift_counts[emp.name] += 1

                    # Update consecutive working days
                    if emp.name in last_work_date and (date - last_work_date[emp.name]).days == 1:
                        consecutive_work_days[emp.name] += 1
                    else:
                        consecutive_work_days[emp.name] = 1
                    last_work_date[emp.name] = date

                # Check if we couldn't fill all positions
                if assigned_count < needed:
                    print(
                        f"  ERROR: Could not fill all {needed} positions for {shift_id} "
                        f"on {date.strftime('%Y-%m-%d')} (skill: {skill}). "
                        f"Only filled {assigned_count}.")

            # Then assign remaining positions to any available employee
            remaining_needed = total_needed - len(assigned_this_shift)
            if remaining_needed > 0:
                # Get valid candidates (any skill) in tiers
                tier1, tier2, tier3 = get_valid_candidates_tiered(
                    employees_available, week_start, shift_hours, assigned_today, date)

                # Sort each tier
                tier1.sort(key=lambda emp: create_sort_key_for_employee(emp, week_start, shift_hours))
                tier2.sort(key=lambda emp: create_sort_key_for_employee(emp, week_start, shift_hours))
                tier3.sort(key=lambda emp: create_sort_key_for_employee(emp, week_start, shift_hours))

                # Try to fill from tier 1 first, then tier 2, then tier 3
                candidates = tier1 + tier2 + tier3

                assigned_count = 0
                used_tier = 1
                for i, emp in enumerate(candidates):
                    if assigned_count >= remaining_needed:
                        break

                    # Track which tier we're using
                    if i >= len(tier1) + len(tier2):
                        if used_tier < 3:
                            print(f"  WARNING: Using emergency tier for {shift_id} on {date.strftime('%Y-%m-%d')}")
                            used_tier = 3
                    elif i >= len(tier1):
                        used_tier = 2

                    shift_assignments[emp.name][date] = shift_id
                    assigned_today.add(emp.name)
                    assigned_count += 1

                    # Update tracking
                    hours_per_week[(emp.name, week_start)] += shift_hours
                    total_hours[emp.name] += shift_hours
                    shift_counts[emp.name] += 1

                    # Update consecutive working days
                    if emp.name in last_work_date and (date - last_work_date[emp.name]).days == 1:
                        consecutive_work_days[emp.name] += 1
                    else:
                        consecutive_work_days[emp.name] = 1
                    last_work_date[emp.name] = date

                # Check if we couldn't fill all positions
                if assigned_count < remaining_needed:
                    print(
                        f"  ERROR: Could not fill all {remaining_needed} remaining positions "
                        f"for {shift_id} on {date.strftime('%Y-%m-%d')}. "
                        f"Only filled {assigned_count}.")

    # Post-processing: Rebalance shifts for better fairness
    print("\nRebalancing shifts for improved fairness...")
    shift_assignments = rebalance_shift_assignments(
        shift_assignments, employees, vacation_dates_by_employee,
        coverage_weekday, coverage_weekend, dates, shifts, employee_by_name
    )

    # Recalculate statistics after rebalancing
    shift_counts = defaultdict(int)
    total_hours = defaultdict(float)
    for emp_name, assignments in shift_assignments.items():
        for date, shift_id in assignments.items():
            shift_counts[emp_name] += 1
            total_hours[emp_name] += calculate_shift_hours(shift_id, shifts)

    # Print fairness statistics
    print("\nShift distribution statistics:")
    working_employees = [emp for emp in employees if shift_counts[emp.name] > 0]
    if working_employees:
        shift_count_list = [shift_counts[emp.name] for emp in working_employees]
        hours_list = [total_hours[emp.name] for emp in working_employees]

        # Verify lists are not empty before calculating statistics
        if shift_count_list and hours_list:
            print(f"  Employees with shifts: {len(working_employees)}")
            print(f"  Shift count range: {min(shift_count_list)} - {max(shift_count_list)} shifts")
            print(f"  Average shifts per working employee: {sum(shift_count_list)/len(shift_count_list):.1f}")
            print(f"  Total hours range: {min(hours_list):.1f} - {max(hours_list):.1f} hours")
            print(f"  Average hours per working employee: {sum(hours_list)/len(hours_list):.1f}")
        else:
            print("  No shifts assigned yet")

    return shift_assignments


def export_schedule_to_excel(
    vacation_schedule: Dict[str, List[datetime]],
    employees: List[Employee],
    coverage_weekday: List[CoverageRequirement],
    coverage_weekend: List[CoverageRequirement],
    shifts: Dict[str, 'Shift'],
    start_date: datetime,
    num_weeks: int,
    filename: str = "vacation_schedule.xlsx"
) -> Optional[str]:
    """Export vacation schedule to Excel file with visual calendar view and shift coverage.

    Args:
        vacation_schedule: Dict mapping employee name -> list of vacation dates
        employees: List of all employees
        coverage_weekday: Coverage requirements for weekdays
        coverage_weekend: Coverage requirements for weekends
        shifts: Dict mapping shift name to Shift object
        start_date: Start date of vacation period
        num_weeks: Number of weeks in the period
        filename: Output filename (default: vacation_schedule.xlsx)

    Returns:
        Path to the generated file, or None if openpyxl is not available
    """
    if not OPENPYXL_AVAILABLE:
        print("\nWarning: openpyxl not available. Cannot generate Excel file.")
        print("Install with: pip install openpyxl")
        return None

    # Create workbook
    wb = Workbook()

    # Create vacation schedule sheet
    ws_vacation = wb.active
    ws_vacation.title = "Vacation Schedule"

    # Generate all dates in the period
    dates = []
    current = start_date
    for _ in range(num_weeks * 7):
        dates.append(current)
        current += timedelta(days=1)

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    vacation_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    working_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
    weekend_header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    # Define shift category color fills
    day_shift_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")     # Light yellow
    evening_shift_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")  # Light blue
    night_shift_fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")   # Light purple/thistle
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # Write header row (dates) for vacation schedule
    ws_vacation.cell(1, 1, "Employee").fill = header_fill
    ws_vacation.cell(1, 1).font = header_font
    ws_vacation.cell(1, 1).alignment = center_align
    ws_vacation.cell(1, 1).border = border
    ws_vacation.column_dimensions['A'].width = 20

    for col_idx, date in enumerate(dates, start=2):
        cell = ws_vacation.cell(1, col_idx)
        day_name = date.strftime('%a')
        date_str = date.strftime('%d/%m')
        cell.value = f"{day_name}\n{date_str}"
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

        # Different color for weekends
        if date.weekday() >= 5:
            cell.fill = weekend_header_fill
        else:
            cell.fill = header_fill

        ws_vacation.column_dimensions[cell.column_letter].width = 8

    # Calculate week boundaries for hour tracking columns
    weeks_info = []
    current_week_start = dates[0] - timedelta(days=dates[0].weekday())  # Get Monday of first week
    for week_num in range(num_weeks):
        week_start = current_week_start + timedelta(weeks=week_num)
        week_end = week_start + timedelta(days=6)
        weeks_info.append((week_start, week_end))

    # Add weekly hours columns
    col_offset = len(dates) + 2
    for week_idx, (week_start, week_end) in enumerate(weeks_info):
        week_col = col_offset + week_idx
        cell = ws_vacation.cell(1, week_col)
        cell.value = f"Week {week_idx + 1}\nHours"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws_vacation.column_dimensions[cell.column_letter].width = 10

    # Add weekly workload percentage columns (% of target hours)
    target_pct_col_start = col_offset + num_weeks
    for week_idx in range(num_weeks):
        week_col = target_pct_col_start + week_idx
        cell = ws_vacation.cell(1, week_col)
        cell.value = f"Week {week_idx + 1}\n% Target"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws_vacation.column_dimensions[cell.column_letter].width = 10

    # Add weekly workload percentage columns (% of max hours)
    max_pct_col_start = target_pct_col_start + num_weeks
    for week_idx in range(num_weeks):
        week_col = max_pct_col_start + week_idx
        cell = ws_vacation.cell(1, week_col)
        cell.value = f"Week {week_idx + 1}\n% Max"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws_vacation.column_dimensions[cell.column_letter].width = 10

    # Add "Total Vacation" and "Total Hours" columns
    total_vacation_col = max_pct_col_start + num_weeks
    ws_vacation.cell(1, total_vacation_col, "Total\nVacation").fill = header_fill
    ws_vacation.cell(1, total_vacation_col).font = header_font
    ws_vacation.cell(1, total_vacation_col).alignment = center_align
    ws_vacation.cell(1, total_vacation_col).border = border
    ws_vacation.column_dimensions[ws_vacation.cell(1, total_vacation_col).column_letter].width = 10

    total_hours_col = total_vacation_col + 1
    ws_vacation.cell(1, total_hours_col, "Total\nHours").fill = header_fill
    ws_vacation.cell(1, total_hours_col).font = header_font
    ws_vacation.cell(1, total_hours_col).alignment = center_align
    ws_vacation.cell(1, total_hours_col).border = border
    ws_vacation.column_dimensions[ws_vacation.cell(1, total_hours_col).column_letter].width = 10

    # Add "% Target" column (percentage of accumulated target hours)
    total_pct_target_col = total_hours_col + 1
    ws_vacation.cell(1, total_pct_target_col, "% Target").fill = header_fill
    ws_vacation.cell(1, total_pct_target_col).font = header_font
    ws_vacation.cell(1, total_pct_target_col).alignment = center_align
    ws_vacation.cell(1, total_pct_target_col).border = border
    ws_vacation.column_dimensions[ws_vacation.cell(1, total_pct_target_col).column_letter].width = 10

    # Convert vacation_schedule to set of dates for faster lookup
    vacation_dates_by_employee = {
        name: set(dates) for name, dates in vacation_schedule.items()
    }

    # Assign shifts to employees
    shift_assignments = assign_shifts_to_employees(
        employees, vacation_schedule, coverage_weekday, coverage_weekend, dates, shifts
    )

    # Write employee rows
    for row_idx, emp in enumerate(sorted(employees, key=lambda e: e.name), start=2):
        # Employee name
        name_cell = ws_vacation.cell(row_idx, 1, emp.name)
        name_cell.border = border
        name_cell.alignment = Alignment(vertical='center')

        vacation_count = 0
        total_work_hours = 0.0

        # Track hours per week for this employee
        hours_by_week = {week_idx: 0.0 for week_idx in range(num_weeks)}

        # Mark vacation days or shift assignments
        for col_idx, date in enumerate(dates, start=2):
            cell = ws_vacation.cell(row_idx, col_idx)
            cell.border = border
            cell.alignment = center_align

            if date in vacation_dates_by_employee.get(emp.name, set()):
                cell.value = "V"
                cell.fill = vacation_fill
                cell.font = Font(bold=True)
                vacation_count += 1
            else:
                # Show shift assignment
                assigned_shift = shift_assignments[emp.name].get(date, "")
                cell.value = assigned_shift
                cell.fill = working_fill  # Default fill
                if assigned_shift:
                    cell.font = Font(size=9)

                    # Calculate hours for this shift and apply category-based coloring
                    if assigned_shift in shifts:
                        shift_obj = shifts[assigned_shift]

                        # Apply color based on shift category
                        if shift_obj.category == "Day":
                            cell.fill = day_shift_fill
                        elif shift_obj.category == "Evening":
                            cell.fill = evening_shift_fill
                        elif shift_obj.category == "Night":
                            cell.fill = night_shift_fill
                        else:
                            cell.fill = working_fill  # Fallback

                        try:
                            start_time = datetime.strptime(shift_obj.start, "%H:%M")
                            end_time = datetime.strptime(shift_obj.end, "%H:%M")
                            shift_hours = (end_time - start_time).total_seconds() / 3600.0
                            if shift_hours < 0:
                                shift_hours += 24  # Handle overnight shifts
                        except (ValueError, IndexError, AttributeError):
                            shift_hours = DEFAULT_SHIFT_HOURS
                    else:
                        shift_hours = DEFAULT_SHIFT_HOURS

                    # Determine which week this date belongs to
                    for week_idx, (week_start, week_end) in enumerate(weeks_info):
                        if week_start <= date <= week_end:
                            hours_by_week[week_idx] += shift_hours
                            break

                    total_work_hours += shift_hours

        # Write weekly hours
        for week_idx in range(num_weeks):
            week_col = col_offset + week_idx
            hours = hours_by_week[week_idx]
            cell = ws_vacation.cell(row_idx, week_col, round(hours, 1) if hours > 0 else "")
            cell.border = border
            cell.alignment = center_align
            if hours > 0:
                cell.font = Font(bold=True)

        # Write weekly workload percentage (% of target hours)
        for week_idx in range(num_weeks):
            week_col = target_pct_col_start + week_idx
            hours = hours_by_week[week_idx]
            if hours > 0 and emp.weekly_target_hours > 0:
                pct = (hours / emp.weekly_target_hours) * 100
                cell = ws_vacation.cell(row_idx, week_col, f"{round(pct, 1)}%")
                cell.border = border
                cell.alignment = center_align
                # Color code based on percentage
                if pct > 100:
                    cell.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")  # Light red
                    cell.font = Font(bold=True, color="CC0000")
                elif pct >= 90:
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light yellow
                    cell.font = Font(bold=True)
                else:
                    cell.font = Font(bold=True)
            else:
                cell = ws_vacation.cell(row_idx, week_col, "")
                cell.border = border
                cell.alignment = center_align

        # Write weekly workload percentage (% of max hours)
        for week_idx in range(num_weeks):
            week_col = max_pct_col_start + week_idx
            hours = hours_by_week[week_idx]
            if hours > 0 and emp.max_hours_per_week > 0:
                pct = (hours / emp.max_hours_per_week) * 100
                cell = ws_vacation.cell(row_idx, week_col, f"{round(pct, 1)}%")
                cell.border = border
                cell.alignment = center_align
                # Color code based on percentage
                if pct > 100:
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
                    cell.font = Font(bold=True, color="FFFFFF")
                elif pct >= 95:
                    cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Orange
                    cell.font = Font(bold=True)
                elif pct >= 85:
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light yellow
                    cell.font = Font(bold=True)
                else:
                    cell.font = Font(bold=True)
            else:
                cell = ws_vacation.cell(row_idx, week_col, "")
                cell.border = border
                cell.alignment = center_align

        # Total vacation days
        vacation_cell = ws_vacation.cell(row_idx, total_vacation_col, vacation_count)
        vacation_cell.border = border
        vacation_cell.alignment = center_align
        vacation_cell.font = Font(bold=True)

        # Total working hours
        hours_cell = ws_vacation.cell(row_idx, total_hours_col, round(
            total_work_hours, 1) if total_work_hours > 0 else "")
        hours_cell.border = border
        hours_cell.alignment = center_align
        if total_work_hours > 0:
            hours_cell.font = Font(bold=True)

        # Total % Target (percentage of accumulated target hours)
        # Calculate accumulated target hours based on working days
        working_days = len(dates) - vacation_count  # Total days - vacation days
        if working_days > 0 and emp.weekly_target_hours > 0:
            # Accumulated target = (working_days / 7) * weekly_target_hours
            accumulated_target_hours = (working_days / 7.0) * emp.weekly_target_hours
            pct_target = (total_work_hours / accumulated_target_hours) * 100
            pct_cell = ws_vacation.cell(row_idx, total_pct_target_col, f"{round(pct_target, 1)}%")
            pct_cell.border = border
            pct_cell.alignment = center_align
            # Color code based on percentage
            if pct_target > 100:
                pct_cell.fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")  # Light red
                pct_cell.font = Font(bold=True, color="CC0000")
            elif pct_target >= 90:
                pct_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light yellow
                pct_cell.font = Font(bold=True)
            else:
                pct_cell.font = Font(bold=True)
        else:
            pct_cell = ws_vacation.cell(row_idx, total_pct_target_col, "")
            pct_cell.border = border
            pct_cell.alignment = center_align

    # Add summary row
    summary_row = len(employees) + 3
    ws_vacation.cell(summary_row, 1, "Employees on vacation:").font = Font(bold=True)
    ws_vacation.cell(summary_row, 1).alignment = Alignment(vertical='center')

    for col_idx, date in enumerate(dates, start=2):
        count = sum(1 for emp in employees
                    if date in vacation_dates_by_employee.get(emp.name, set()))
        cell = ws_vacation.cell(summary_row, col_idx, count)
        cell.alignment = center_align
        cell.font = Font(bold=True)
        cell.border = border

        # Color code based on count
        if count > 0:
            # Gradient from light green (90EE90) to darker green
            intensity = min(count / 10.0, 1.0)
            # Calculate green component: 238 (0xEE) down to 144 (0x90)
            green_val = int(238 - (238 - 144) * intensity)
            green_val = max(144, min(238, green_val))  # Clamp to valid range
            hex_color = f"90{green_val:02X}90"
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # Add coverage percentage row
    coverage_row = summary_row + 1
    ws_vacation.cell(coverage_row, 1, "Coverage %:").font = Font(bold=True)
    ws_vacation.cell(coverage_row, 1).alignment = Alignment(vertical='center')

    for col_idx, date in enumerate(dates, start=2):
        is_weekend = date.weekday() >= 5
        requirements = coverage_weekend if is_weekend else coverage_weekday

        # Calculate total required positions for this day
        total_required = sum(req.required for req in requirements)

        # Count actual assignments - each employee should only be counted once
        employees_assigned = set()
        for emp in employees:
            assigned_shift = shift_assignments[emp.name].get(date, "")
            if assigned_shift:
                employees_assigned.add(emp.name)

        total_assigned = len(employees_assigned)

        # Calculate coverage percentage
        if total_required > 0:
            coverage_pct = (total_assigned / total_required) * 100
        else:
            coverage_pct = 100.0

        cell = ws_vacation.cell(coverage_row, col_idx, f"{coverage_pct:.0f}%")
        cell.alignment = center_align
        cell.font = Font(bold=True)
        cell.border = border

        # Color code based on coverage percentage
        if coverage_pct < 80:
            # Red - poor coverage
            cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            cell.font = Font(bold=True, color="990000")
        elif coverage_pct < 100:
            # Yellow - partial coverage
            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        else:
            # Green - full or over coverage
            cell.fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    # Freeze panes
    ws_vacation.freeze_panes = 'B2'

    # ==================================================================
    # CREATE SHIFT COVERAGE SHEET
    # ==================================================================
    ws_coverage = wb.create_sheet("Shift Coverage")

    # Define additional styles for coverage sheet
    shift_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")

    # Write header row
    ws_coverage.cell(1, 1, "Date").fill = header_fill
    ws_coverage.cell(1, 1).font = header_font
    ws_coverage.cell(1, 1).alignment = center_align
    ws_coverage.cell(1, 1).border = border
    ws_coverage.column_dimensions['A'].width = 12

    ws_coverage.cell(1, 2, "Day").fill = header_fill
    ws_coverage.cell(1, 2).font = header_font
    ws_coverage.cell(1, 2).alignment = center_align
    ws_coverage.cell(1, 2).border = border
    ws_coverage.column_dimensions['B'].width = 10

    ws_coverage.cell(1, 3, "Shift").fill = header_fill
    ws_coverage.cell(1, 3).font = header_font
    ws_coverage.cell(1, 3).alignment = center_align
    ws_coverage.cell(1, 3).border = border
    ws_coverage.column_dimensions['C'].width = 10

    ws_coverage.cell(1, 4, "Time").fill = header_fill
    ws_coverage.cell(1, 4).font = header_font
    ws_coverage.cell(1, 4).alignment = center_align
    ws_coverage.cell(1, 4).border = border
    ws_coverage.column_dimensions['D'].width = 15

    ws_coverage.cell(1, 5, "Required").fill = header_fill
    ws_coverage.cell(1, 5).font = header_font
    ws_coverage.cell(1, 5).alignment = center_align
    ws_coverage.cell(1, 5).border = border
    ws_coverage.column_dimensions['E'].width = 10

    ws_coverage.cell(1, 6, "Skill").fill = header_fill
    ws_coverage.cell(1, 6).font = header_font
    ws_coverage.cell(1, 6).alignment = center_align
    ws_coverage.cell(1, 6).border = border
    ws_coverage.column_dimensions['F'].width = 12

    ws_coverage.cell(1, 7, "Available").fill = header_fill
    ws_coverage.cell(1, 7).font = header_font
    ws_coverage.cell(1, 7).alignment = center_align
    ws_coverage.cell(1, 7).border = border
    ws_coverage.column_dimensions['G'].width = 10

    # Write coverage data
    row = 2
    for date in dates:
        is_weekend = date.weekday() >= 5
        requirements = coverage_weekend if is_weekend else coverage_weekday
        day_name = date.strftime('%A')
        date_str = date.strftime('%Y-%m-%d')

        # Get employees available on this date (not on vacation)
        employees_available = [emp for emp in employees
                               if date not in vacation_dates_by_employee.get(emp.name, set())]

        # Group requirements by shift
        shift_reqs = defaultdict(list)
        for req in requirements:
            shift_reqs[req.shift_id].append(req)

        # Write each shift's requirements
        for shift_id in sorted(shift_reqs.keys()):
            reqs = shift_reqs[shift_id]
            shift = shifts.get(shift_id)
            shift_time = f"{shift.start}-{shift.end}" if shift else "N/A"

            for req in reqs:
                ws_coverage.cell(row, 1, date_str).border = border
                ws_coverage.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')

                ws_coverage.cell(row, 2, day_name).border = border
                ws_coverage.cell(row, 2).alignment = center_align
                if is_weekend:
                    ws_coverage.cell(row, 2).fill = weekend_header_fill
                    ws_coverage.cell(row, 2).font = Font(color="FFFFFF", bold=True)

                ws_coverage.cell(row, 3, shift_id).border = border
                ws_coverage.cell(row, 3).alignment = center_align
                ws_coverage.cell(row, 3).fill = shift_fill

                ws_coverage.cell(row, 4, shift_time).border = border
                ws_coverage.cell(row, 4).alignment = center_align

                ws_coverage.cell(row, 5, req.required).border = border
                ws_coverage.cell(row, 5).alignment = center_align
                ws_coverage.cell(row, 5).font = Font(bold=True)

                skill_text = req.required_skill if req.required_skill != "None" else "Any"
                ws_coverage.cell(row, 6, skill_text).border = border
                ws_coverage.cell(row, 6).alignment = center_align

                # Count available employees with required skill
                if req.required_skill == "None":
                    available_count = len(employees_available)
                else:
                    available_count = sum(1 for emp in employees_available
                                          if req.required_skill in emp.skills)

                ws_coverage.cell(row, 7, available_count).border = border
                ws_coverage.cell(row, 7).alignment = center_align

                # Color code based on coverage adequacy
                if available_count < req.required:
                    # Red - insufficient coverage
                    ws_coverage.cell(row, 7).fill = PatternFill(
                        start_color="FF9999", end_color="FF9999", fill_type="solid")
                    ws_coverage.cell(row, 7).font = Font(bold=True, color="990000")
                elif available_count == req.required:
                    # Yellow - exact coverage
                    ws_coverage.cell(row, 7).fill = PatternFill(
                        start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    ws_coverage.cell(row, 7).font = Font(bold=True)
                else:
                    # Green - good coverage
                    ws_coverage.cell(row, 7).fill = PatternFill(
                        start_color="99FF99", end_color="99FF99", fill_type="solid")

                row += 1

    # Freeze panes on coverage sheet
    ws_coverage.freeze_panes = 'A2'

    # Save workbook
    wb.save(filename)
    return filename


def main():
    """Main entry point."""
    employees_file = "employees.csv"
    shifts_file = "shifts.csv"
    coverage_file = "coverage.csv"

    # Vacation period parameters
    # Summer vacation: Week 27 (June 29) to end of Week 31 (August 2)
    start_year = 2026
    start_month = 6
    start_day = 29
    num_weeks = 5  # 5-week range (35 days)
    target_days = 21

    # Parse command line arguments
    if len(sys.argv) > 1:
        employees_file = sys.argv[1]
    if len(sys.argv) > 2:
        coverage_file = sys.argv[2]
    if len(sys.argv) > 3:
        start_date_str = sys.argv[3]
        try:
            parts = start_date_str.split('-')
            start_year = int(parts[0])
            start_month = int(parts[1])
            start_day = int(parts[2])
        except (ValueError, IndexError):
            print(f"Warning: Invalid start date '{start_date_str}', using default", file=sys.stderr)
    if len(sys.argv) > 4:
        num_weeks = int(sys.argv[4])
    if len(sys.argv) > 5:
        target_days = int(sys.argv[5])

    print(f"Loading employees from: {employees_file}")
    employees = load_employees(employees_file)
    print(f"  Loaded {len(employees)} employees")

    print(f"Loading shift definitions from: {shifts_file}")
    shifts = load_shifts(shifts_file)
    print(f"  Loaded {len(shifts)} shift types")

    print(f"Loading coverage requirements from: {coverage_file}")
    coverage = load_coverage(coverage_file)
    print(f"  Loaded {len(coverage)} coverage requirements")

    # Organize coverage by day type
    coverage_by_type = get_coverage_requirements_by_type(coverage)
    coverage_weekday = coverage_by_type.get('Weekday', [])
    coverage_weekend = coverage_by_type.get('Weekend', [])

    start_date = datetime(start_year, start_month, start_day)

    print("\nOptimizing vacation schedule:")
    print(f"  Start date: {start_date.date()}")
    print(f"  Duration: {num_weeks} weeks ({num_weeks * 7} days)")
    print(f"  Target: {target_days} days per employee")

    # Optimize vacation schedule
    vacation_schedule = optimize_vacation_schedule(
        employees,
        coverage_weekday,
        coverage_weekend,
        start_date,
        num_weeks,
        target_days
    )

    # Print results
    print_vacation_results(vacation_schedule, employees, num_weeks, target_days)

    # Export to Excel
    print("\nGenerating Excel schedule...")
    excel_file = export_schedule_to_excel(
        vacation_schedule,
        employees,
        coverage_weekday,
        coverage_weekend,
        shifts,
        start_date,
        num_weeks,
        "vacation_schedule.xlsx"
    )

    if excel_file:
        print(f"✓ Excel schedule saved to: {excel_file}")
    else:
        print("✗ Could not generate Excel file (openpyxl not available)")


if __name__ == "__main__":
    main()
